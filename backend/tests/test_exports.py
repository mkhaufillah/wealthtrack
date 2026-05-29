"""Tests for /api/v1/exports endpoints."""

import io

import openpyxl
from httpx import AsyncClient


class TestExportYearly:
    async def test_export_returns_xlsx(self, client: AsyncClient, filla_token: str):
        """GET /exports/yearly?year=2026 returns 200 with xlsx content-type."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2026",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    async def test_export_requires_auth(self, client: AsyncClient):
        """GET /exports/yearly without token returns 401."""
        resp = await client.get("/api/v1/exports/yearly?year=2026")
        assert resp.status_code == 401

    async def test_export_invalid_year(self, client: AsyncClient, filla_token: str):
        """GET /exports/yearly with year out of range returns 422."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=1900",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 422

    async def test_export_xlsx_content_valid(
        self, client: AsyncClient, filla_token: str
    ):
        """Verify the xlsx file can be parsed correctly with openpyxl."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2026",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 200

        # Parse the xlsx
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb is not None

        # Should have 12 sheets (Jan through Dec)
        sheet_names = wb.sheetnames
        expected_months = [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
        ]
        for mo in expected_months:
            assert mo in sheet_names, f"Sheet '{mo}' not found in {sheet_names}"

    async def test_export_xlsx_has_data_rows(
        self, client: AsyncClient, filla_token: str
    ):
        """Sheets with seed transactions should have data rows."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2026",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))

        # Find which month(s) have data (seed data is from recent days so could be in May 2026)
        total_data_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Count non-header rows before the summary section
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # blank row marks start of summary section
                    break
                if row[0] and row[3]:  # has date and amount
                    row_count += 1
            total_data_rows += row_count

        assert total_data_rows >= 1, "Expected at least 1 data row across all sheets"

    async def test_export_xlsx_summary_totals(
        self, client: AsyncClient, filla_token: str
    ):
        """Verify the summary rows (Total Income / Total Expense / Balance) exist."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2026",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))

        found_income = False
        found_expense = False
        found_balance = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if row[0] == "Total Income":
                    found_income = True
                    assert row[3] is not None  # total income is present
                elif row[0] == "Total Expense":
                    found_expense = True
                    assert row[3] is not None  # total expense is present
                elif row[0] == "Balance":
                    found_balance = True
                    assert row[3] is not None  # balance is present

        assert found_income, "No Total Income summary row found"
        assert found_expense, "No Total Expense summary row found"
        assert found_balance, "No Balance summary row found"

    async def test_export_no_transactions(
        self, client: AsyncClient, filla_token: str
    ):
        """Export for a year with no transactions → valid xlsx with empty sheets."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2099",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        # Should still have 12 sheets
        assert len(wb.sheetnames) == 12
        # Each sheet should still have headers but no data rows
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Row 1 is header
            header = [ws.cell(row=1, column=c).value for c in range(1, 8)]
            assert header == ["Date", "Type", "Category", "Amount", "Description", "Note", "Owner"]

    async def test_export_xlsx_content_disposition(
        self, client: AsyncClient, filla_token: str
    ):
        """Response has proper Content-Disposition header."""
        resp = await client.get(
            "/api/v1/exports/yearly?year=2026",
            headers={"Authorization": f"Bearer {filla_token}"},
        )
        assert "Content-Disposition" in resp.headers
        assert "attachment" in resp.headers["content-disposition"]
        assert "wealthtrack_2026.xlsx" in resp.headers["content-disposition"]

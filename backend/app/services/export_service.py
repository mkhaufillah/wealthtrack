"""Export Service — business logic for all export endpoints.

Generates Excel (.xlsx) reports from transaction data.
No FastAPI dependency — works with CursorWrapper (from app.database) directly.

Usage::

    service = ExportService(db)
    buf, filename = await service.export_yearly(year=2026, user_id=1)
    # buf is a BytesIO ready to send as a StreamingResponse
"""

import io
from calendar import month_abbr
from app.database import CursorWrapper


class ExportService:
    """Service for all export/generation operations.

    Instantiate with a CursorWrapper (from ``app.database.get_db``).
    All methods return plain Python types — no FastAPI types.
    """

    def __init__(self, db: CursorWrapper) -> None:
        self.db = db

    # ── Helpers ──────────────────────────────────────────────────────────

    async def _year_transactions(self, year: int, user_id: int) -> dict[int, list]:
        """Fetch all transactions for a given year, grouped by month."""
        cursor = await self.db.execute(
            """SELECT t.id, t.type, t.amount, t.category_name, t.description, t.note,
                      t.date, t.user_id, t.created_at, u.display_name AS owner_name
               FROM transactions t
               LEFT JOIN users u ON t.user_id = u.id
               WHERE t.user_id = ?
                 AND COALESCE(t.date, LEFT(t.created_at::text, 10)) >= ?
                 AND COALESCE(t.date, LEFT(t.created_at::text, 10)) <= ?
               ORDER BY COALESCE(t.date, LEFT(t.created_at::text, 10)) ASC""",
            (user_id, f"{year}-01-01", f"{year}-12-31"),
        )
        rows = await cursor.fetchall()
        grouped: dict[int, list] = {m: [] for m in range(1, 13)}
        for r in rows:
            d = dict(r)
            raw_date = d["date"] or d["created_at"][:10]
            mo = int(raw_date[5:7])
            grouped[mo].append(d)
        return grouped

    # ── Yearly Export ────────────────────────────────────────────────────

    async def export_yearly(self, year: int, user_id: int) -> tuple[io.BytesIO, str]:
        """Generate a yearly .xlsx workbook — one sheet per month.

        Returns a ``(BytesIO, filename)`` tuple. The caller (typically a
        FastAPI router) wraps the buffer in a ``StreamingResponse``.

        Parameters
        ----------
        year : int
            Calendar year to export (2020–2100).
        user_id : int
            ID of the user whose transactions to export.

        Returns
        -------
        tuple[io.BytesIO, str]
            (xlsx buffer, suggested filename like ``"wealthtrack_2026.xlsx"``)
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        grouped = await self._year_transactions(year, user_id)

        wb = openpyxl.Workbook()
        # Remove default sheet — we'll add named ones
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        summary_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

        headers = ["Date", "Type", "Category", "Amount", "Description", "Note", "Owner"]
        col_widths = [14, 10, 20, 16, 30, 20, 16]

        for mo in range(1, 13):
            ws = wb.create_sheet(title=month_abbr[mo])
            rows = grouped[mo]

            # Headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Column widths
            for col_idx, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

            # Data rows
            total_income = 0
            total_expense = 0
            for row_idx, txn in enumerate(rows, 2):
                raw_date = txn["date"] or txn["created_at"][:10]
                ws.cell(row=row_idx, column=1, value=raw_date).border = thin_border
                ws.cell(row=row_idx, column=2, value=txn["type"].capitalize()).border = thin_border
                ws.cell(row=row_idx, column=3, value=txn["category_name"] or "").border = thin_border
                ws.cell(row=row_idx, column=4, value=txn["amount"]).border = thin_border
                ws.cell(row=row_idx, column=5, value=txn["description"] or "").border = thin_border
                ws.cell(row=row_idx, column=6, value=txn["note"] or "").border = thin_border
                ws.cell(row=row_idx, column=7, value=txn["owner_name"] or "").border = thin_border

                if txn["type"] == "income":
                    total_income += txn["amount"]
                else:
                    total_expense += txn["amount"]

            # Summary rows (Total Income / Expense / Balance)
            summary_row = len(rows) + 3  # blank row then summary

            ws.cell(row=summary_row, column=1, value="Total Income").font = summary_font
            ws.cell(row=summary_row, column=4, value=total_income).font = Font(color="2ECC71", bold=True)

            ws.cell(row=summary_row + 1, column=1, value="Total Expense").font = summary_font
            ws.cell(row=summary_row + 1, column=4, value=total_expense).font = Font(color="E94560", bold=True)

            balance = total_income - total_expense
            ws.cell(row=summary_row + 2, column=1, value="Balance").font = summary_font
            bal_cell = ws.cell(row=summary_row + 2, column=4, value=balance)
            bal_cell.font = Font(bold=True, color="0F3460" if balance >= 0 else "E94560")
            bal_cell.number_format = "#,##0"

        # Write to BytesIO
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"wealthtrack_{year}.xlsx"
        return buf, filename
